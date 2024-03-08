from openpyxl import load_workbook
import csv
import os


def extract_sheet_in_excel_file(excel_path=f'{os.getcwd()}/excel_to_csv/loss_function'):
    csv_path = f'{os.getcwd()}/excel_to_csv/csv_loss'
    list_source_file = os.listdir(excel_path)

    for file in list_source_file:
        workbook = load_workbook(f'{excel_path}/{file}')

        # get the names of all the sheet in excel file
        name_sheets = workbook.sheetnames
        for name in name_sheets:
            sheet = workbook[name]  # get sheet by name

            # store sheet in a csv file
            with open(f'{csv_path}/{name}.csv', 'w', newline='') as csv_file:
                writer = csv.writer(csv_file, delimiter=',')
                # writer.writeheader()
                for row in sheet.iter_rows(values_only=True):
                    # print(row)
                    writer.writerow(row)
            # close the file
            csv_file.close()


extract_sheet_in_excel_file()
